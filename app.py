import streamlit as st
import asyncio
import aiohttp
import pandas as pd
from collections import Counter
import re
from datetime import datetime
import random
import time
import io
import os
import json
import logging
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import plotly.express as px
import plotly.graph_objects as go
from wordcloud import WordCloud
import matplotlib.pyplot as plt
from dotenv import load_dotenv

# Загрузка переменных окружения
load_dotenv()

# Настройка логирования
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler('app.log'),
        logging.StreamHandler()
    ]
)

# ===== ИМПОРТ КОНФИГУРАТОРА =====
from config_generator import (
    config, 
    generate_config_from_api, 
    DEFAULT_TECH_CATEGORIES,
    DEFAULT_ACTION_VERBS,
    DEFAULT_CATEGORY_ICONS,
    DEFAULT_CATEGORY_COLORS
)

# ====================== КОНФИГУРАЦИЯ STREAMLIT ======================
st.set_page_config(
    page_title="Аналитический Дашборд (поиск ключевых слов по должности)",
    page_icon="logo.svg",
    layout="wide",
    initial_sidebar_state="expanded"
)

# CSS стили для темной темы
st.markdown("""
<style>
    /* Скрываем служебные элементы Streamlit */
    #MainMenu {visibility: hidden;}
    footer {visibility: hidden;}
    [data-testid="stDecoration"] {display: none !important;}
    [data-testid="stAppDeployButton"] {display: none !important;}

    /* Основные цвета для темной темы */
    .stApp {
        background-color: #0e1117;
    }
    
    .main-header { 
        font-size: 2.5rem; 
        font-weight: bold; 
        color: #00FF88 !important;
        text-shadow: 0 0 10px rgba(0,255,136,0.3);
    }
    
    .success-box { 
        background-color: #1a3a2a; 
        padding: 10px; 
        border-radius: 5px; 
        border-left: 4px solid #00FF88;
        color: #00FF88 !important;
    }
    
    .warning-box { 
        background-color: #3a2a1a; 
        padding: 10px; 
        border-radius: 5px; 
        border-left: 4px solid #ffc107;
        color: #ffc107 !important;
    }
    
    /* Стили для карточек категорий */
    .category-card {
        display: flex;
        align-items: center;
        gap: 12px;
        padding: 12px;
        margin-bottom: 10px;
        background: linear-gradient(135deg, #1a1f2e 0%, #0f1420 100%);
        border-radius: 10px;
        border-left: 5px solid;
        box-shadow: 0 2px 8px rgba(0,0,0,0.3);
        transition: transform 0.2s;
        cursor: pointer;
    }
    .category-card:hover {
        transform: translateX(5px);
    }
    .category-icon {
        font-size: 24px;
    }
    .category-color-box {
        width: 24px;
        height: 24px;
        border-radius: 6px;
        box-shadow: 0 2px 4px rgba(0,0,0,0.2);
        flex-shrink: 0;
    }
    .category-info {
        flex: 1;
    }
    .category-title {
        font-weight: bold;
        color: #e0e0e0 !important;
        font-size: 14px;
    }
    .category-stats {
        color: #a0a0a0 !important;
        font-size: 12px;
    }
    .category-percent {
        margin-left: 8px;
        padding: 2px 8px;
        border-radius: 12px;
        font-weight: bold;
        font-size: 11px;
    }
    
    /* Стили для таблиц */
    .stDataFrame {
        background-color: #1a1f2e;
    }
    .dataframe {
        color: #e0e0e0 !important;
    }
    
    /* Метрики */
    .stMetric {
        background-color: #1a1f2e;
        border-radius: 10px;
        padding: 10px;
    }
    
    /* Информационные сообщения */
    .stAlert {
        background-color: #1a1f2e;
        color: #e0e0e0;
    }
    
    /* Кнопки */
    .stButton button {
        background-color: #00FF88;
        color: #0e1117;
        font-weight: bold;
        border-radius: 8px;
        transition: all 0.3s;
    }
    .stButton button:hover {
        background-color: #00cc6a;
        transform: translateY(-2px);
    }
</style>
""", unsafe_allow_html=True)

# ====================== КОНСТАНТЫ ======================
USER_AGENTS = [
    'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36',
    'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/119.0.0.0 Safari/537.36',
    'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36',
]

COMMON_STOPWORDS = [
    'strong', 'quot', 'nbsp', 'lt', 'gt', 'amp', 'rsquo', 'lsquo', 'mdash', 'br', 'li', 'ul', 'ol',
    'работа', 'работы', 'работе', 'работу', 'рабочий', 'рабочих', 'вакансия', 'компания', 'компании',
    'сотрудник', 'сотрудников', 'офис', 'график', 'зарплата', 'оформление', 'трудоустройство',
    'социальный', 'пакет', 'дмс', 'корпоратив', 'премия', 'больничный', 'отпуск', 'льготы', 'условие',
    'условия', 'требование', 'требования', 'обязанность', 'обязанности', 'возможность', 'умение',
    'знание', 'понимание', 'опыт', 'навык', 'навыки', 'образование', 'высшее', 'средний', 'специальный',
    'сертификат', 'портфолио', 'год', 'года', 'лет', 'месяц', 'день', 'неделя', 'ооо', 'ип', 'зао', 'ао',
    'пао', 'россия', 'москва', 'спб', 'можно', 'нужно', 'должен', 'должны', 'будет', 'быть', 'очень', 'также',
    'кандидат', 'кандидата', 'кандидату', 'позиция', 'должность', 'сфера', 'область', 'направление',
    'команда', 'коллектив', 'проект', 'проекты', 'задача', 'задачи', 'участие', 'мы', 'наш', 'наша', 'наши',
    'вы', 'ваш', 'вам', 'предлагать', 'предлагаем', 'предлагается', 'гарантировать', 'гарантируем',
    'искать', 'ищем', 'рассматривать', 'рассматриваем', 'приглашать', 'приглашаем', 'ждать', 'ждем', 'ждём',
    'приветствоваться', 'приветствуется', 'плюс', 'будет плюсом', 'преимущество', 'преимуществом',
    'желательно', 'обязательно', 'необходимо', 'важно', 'ответственность', 'ответственный',
    'коммуникабельность', 'коммуникабельный', 'стрессоустойчивость', 'стрессоустойчивый', 'обучаемость',
    'инициативность', 'исполнительность', 'внимательность', 'активность', 'нацеленность', 'результат',
    'выполнение', 'выполнять', 'обеспечение', 'обеспечивать', 'организация', 'организовывать', 'контроль',
    'контролировать', 'ведение', 'вести', 'разработка', 'разрабатывать', 'реализация', 'реализовывать',
    'поддержка', 'поддерживать', 'сопровождение', 'сопровождать', 'срок', 'сроки', 'время', 'период',
    'ежедневно', 'ежемесячно', 'ежегодно', 'полный', 'неполный', 'занятость', 'частичный', 'удаленно',
    'удалённо', 'офлайн', 'онлайн', 'гибкий', 'гибкость', 'сменный', 'вахта', 'тк', 'тк рф', 'трудовой',
    'кодекс', 'официальный', 'официально', 'договор', 'испытательный', 'испытательный срок', 'доход',
    'оклад', 'бонус', 'бонусы', 'kpi', 'мотивация', 'компенсация', 'возмещение', 'дружный', 'дружная',
    'молодой', 'молодая', 'развивающийся', 'развивающаяся', 'стабильный', 'стабильная', 'крупный',
    'крупная', 'успешный', 'успешная', 'лидирующий', 'амбициозный', 'динамичный', 'обучение', 'обучаться',
    'развитие', 'карьерный', 'карьера', 'рост', 'город', 'регион', 'локация', 'релокация', 'базовый',
    'продвинутый', 'уверенный', 'желание', 'готовность', 'способность', 'уметь', 'team', 'player', 'skills',
    'skill', 'experience', 'knowledge', 'ability', 'responsibility', 'requirements', 'duties', 'tasks',
    'для', 'или', 'команде', 'заработная', 'плата', 'разработки', 'без', 'официальное', 'что', 'роста', 
    'дня', 'после', 'как', 'предстоит', 'если', 'чем', 'очень', 'также', 'более', 'менее', 'почти', 
    'около', 'примерно', 'действительно', 'конечно', 'вообще', 'просто', 'совсем', 'абсолютно', 
    'относительно', 'этот', 'эта', 'это', 'эти', 'того', 'тому', 'тем', 'этом', 'который', 'которая', 
    'которое', 'которые', 'такой', 'такая', 'такое', 'такие', 'весь', 'вся', 'все', 'всё', 'всех',
    'итак', 'таким', 'образом', 'следовательно', 'во-первых', 'во-вторых', 'наконец', 'кстати', 
    'например', 'вообще-то', 'являться', 'является', 'быть', 'есть', 'стать', 'становиться',
    'находиться', 'находится', 'иметь', 'имеет', 'обладать', 'обладает',
    'a', 'b', 'c', 'd', 'e', 'f', 'g', 'h', 'i', 'j', 'k', 'l', 'm', 
    'n', 'o', 'p', 'q', 'r', 's', 't', 'u', 'v', 'w', 'x', 'y', 'z'
]

RATE_LIMIT_FILE = "ip_generation_limits.json"
MAX_GENERATIONS_PER_DAY = 2

# ====================== ФУНКЦИИ ======================
def get_random_user_agent():
    return random.choice(USER_AGENTS)

def get_client_ip():
    """Пробует получить IP клиента из заголовков (через прокси/напрямую)."""
    try:
        headers = {}
        if hasattr(st, "context") and hasattr(st.context, "headers"):
            headers = dict(st.context.headers)

        xff = headers.get("X-Forwarded-For") or headers.get("x-forwarded-for")
        if xff:
            return xff.split(",")[0].strip()

        x_real_ip = headers.get("X-Real-IP") or headers.get("x-real-ip")
        if x_real_ip:
            return x_real_ip.strip()
    except Exception:
        pass
    return None

def _load_ip_rate_limits():
    if not os.path.exists(RATE_LIMIT_FILE):
        return {}
    try:
        with open(RATE_LIMIT_FILE, "r", encoding="utf-8") as f:
            data = json.load(f)
            return data if isinstance(data, dict) else {}
    except Exception:
        return {}

def _save_ip_rate_limits(data):
    try:
        with open(RATE_LIMIT_FILE, "w", encoding="utf-8") as f:
            json.dump(data, f, ensure_ascii=False, indent=2)
    except Exception as e:
        logging.error(f"Не удалось сохранить лимиты IP: {e}")

def get_generation_usage_today(client_ip):
    """Возвращает (использовано_сегодня, осталось_сегодня) для IP."""
    if not client_ip:
        return 0, MAX_GENERATIONS_PER_DAY

    today = datetime.now().date().isoformat()
    data = _load_ip_rate_limits()
    record = data.get(client_ip)

    # Обратная совместимость со старым форматом: {"ip": "YYYY-MM-DD"}
    if isinstance(record, str):
        used_today = 1 if record == today else 0
    elif isinstance(record, dict):
        if record.get("date") == today:
            used_today = int(record.get("count", 0))
        else:
            used_today = 0
    else:
        used_today = 0

    used_today = max(0, min(used_today, MAX_GENERATIONS_PER_DAY))
    remaining = MAX_GENERATIONS_PER_DAY - used_today
    return used_today, remaining

def can_generate_config_today(client_ip):
    """True, если для IP не исчерпан дневной лимит генераций."""
    _, remaining = get_generation_usage_today(client_ip)
    return remaining > 0

def mark_config_generated_today(client_ip):
    if not client_ip:
        return
    today = datetime.now().date().isoformat()
    data = _load_ip_rate_limits()
    used_today, _ = get_generation_usage_today(client_ip)
    new_count = min(used_today + 1, MAX_GENERATIONS_PER_DAY)
    data[client_ip] = {"date": today, "count": new_count}
    _save_ip_rate_limits(data)

def clean_keywords(keywords, job_titles):
    """Очистка ключевых слов от названий должностей"""
    cleaned = []
    job_phrases = set()
    
    for job in job_titles:
        job_lower = job.lower().strip()
        job_phrases.add(job_lower)
        words = re.findall(r'[a-zа-яё]+', job_lower)
        for word in words:
            if len(word) <= 3 or word in ['менеджер', 'developer', 'специалист', 'инженер', 'аналитик']:
                job_phrases.add(word)
    
    for kw in keywords:
        kw_lower = kw.lower().strip()
        is_job_title = False
        
        if kw_lower in job_phrases:
            is_job_title = True
        
        if kw_lower in ['категорийный', 'менеджер', 'developer', 'разработчик', 'специалист', 
                        'инженер', 'аналитик', 'администратор', 'архитектор', 'дизайнер',
                        'руководитель', 'лид', 'тимлид', 'teamlead', 'team lead']:
            if kw_lower not in config.all_tech_terms:
                is_job_title = True
        
        if not is_job_title:
            cleaned.append(kw)
    
    return cleaned

def extract_keywords(text, job_titles):
    if not text:
        return []
    text = text.lower()
    keywords = set()
    
    for term in config.all_tech_terms:
        if term in text:
            keywords.add(term)
    
    for action in config.action_verbs:
        if action in text:
            keywords.add(action)
    
    words = re.findall(r'\b[a-zа-яё]{4,25}\b', text)
    for word in words:
        if word not in COMMON_STOPWORDS:
            keywords.add(word)
    
    keywords_list = clean_keywords(list(keywords), job_titles)
    return keywords_list

def calculate_rating(frequency, max_frequency):
    if max_frequency == 0 or frequency == 0:
        return 1, "★", 0.0
    
    ratio = frequency / max_frequency
    
    if ratio >= 0.8:
        stars = 5
    elif ratio >= 0.6:
        stars = 4
    elif ratio >= 0.4:
        stars = 3
    elif ratio >= 0.2:
        stars = 2
    else:
        stars = 1
    
    star_text = "★" * stars
    percentage = ratio * 100
    return stars, star_text, round(percentage, 1)

def render_stars(stars):
    stars = int(stars) if isinstance(stars, (int, float, str)) else 1
    colors = {
        5: '#FFD700', 4: '#FFA500', 3: '#FFD966', 
        2: '#B4C6E7', 1: '#D3D3D3'
    }
    star_html = ''
    for i in range(1, 6):
        if i <= stars:
            color = colors.get(stars, '#D3D3D3')
            star_html += f'<span style="color: {color}; font-size: 16px;">★</span>'
        else:
            star_html += '<span style="color: #E0E0E0; font-size: 16px;">★</span>'
    return star_html

def create_excel_with_formatting(dataframes, total_vacancies):
    output = io.BytesIO()
    
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        for sheet_name, df in dataframes.items():
            df.to_excel(writer, sheet_name=sheet_name, index=False)
    
    output.seek(0)
    wb = load_workbook(output)
    star_colors = {5: "FFD700", 4: "FFA500", 3: "FFD966", 2: "B4C6E7", 1: "D3D3D3"}
    
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill(start_color="2E75B6", end_color="2E75B6", fill_type="solid")
        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )
        
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = thin_border
        
        rating_col = None
        for col_idx, cell in enumerate(ws[1], 1):
            if cell.value and "Рейтинг" in str(cell.value):
                rating_col = col_idx
                break
        
        if rating_col:
            for row_idx in range(2, ws.max_row + 1):
                rating_cell = ws.cell(row=row_idx, column=rating_col)
                if rating_cell.value and isinstance(rating_cell.value, str):
                    stars_count = rating_cell.value.count('★')
                    if stars_count in star_colors:
                        rating_cell.font = Font(color=star_colors[stars_count], bold=True, size=12)
                        rating_cell.alignment = Alignment(horizontal="center", vertical="center")
        
        for col in ws.columns:
            max_length = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            ws.column_dimensions[col_letter].width = min(max_length + 3, 50)
        
        ws.insert_rows(1)
        info_cell = ws.cell(row=1, column=1)
        info_cell.value = f"📊 Анализ {total_vacancies} вакансий | ★★★★★=80%+ | ★★★★=60-80% | ★★★=40-60% | ★★=20-40% | ★=<20%"
        info_cell.font = Font(bold=True, italic=True, size=9, color="2E75B6")
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ws.max_column)
    
    output.seek(0)
    return output

# ====================== АСИНХРОННЫЙ ПАРСЕР ======================
class AsyncHHParser:
    def __init__(self, progress_bar, status_text, log_container, job_titles):
        self.progress_bar = progress_bar
        self.status_text = status_text
        self.log_container = log_container
        self.job_titles = job_titles
        self.all_keywords = []
        self.processed = 0
        self.total = 0
        self.is_running = True
        self.session = None
        
    def log(self, message):
        timestamp = datetime.now().strftime("%H:%M:%S")
        if self.log_container:
            self.log_container.text(f"[{timestamp}] {message}")
        logging.info(message)
    
    async def init_session(self):
        self.session = aiohttp.ClientSession(
            headers={'User-Agent': get_random_user_agent()},
            timeout=aiohttp.ClientTimeout(total=10)
        )
    
    async def close_session(self):
        if self.session:
            await self.session.close()
    
    async def get_vacancy_ids(self, job_title, limit):
        ids = []
        per_page = 100
        pages_needed = min(20, (limit // per_page) + 2)
        self.log(f"🔍 Поиск: '{job_title}' (цель: {limit})")
        
        for page in range(pages_needed):
            if not self.is_running:
                break
            url = "https://api.hh.ru/vacancies"
            params = {'text': job_title, 'area': 113, 'page': page, 'per_page': per_page}
            
            try:
                async with self.session.get(url, params=params) as response:
                    if response.status == 200:
                        data = await response.json()
                        items = data.get('items', [])
                        ids.extend([item['id'] for item in items])
                        self.log(f"  Стр. {page+1}: +{len(items)} ID (всего: {len(ids)})")
                        if len(items) < per_page:
                            break
                        await asyncio.sleep(0.2)
            except Exception as e:
                self.log(f"  ⚠ Ошибка: {e}")
                break
        return ids[:limit]
    
    async def get_vacancy_details(self, vacancy_id):
        if not self.is_running:
            return None
        await asyncio.sleep(random.uniform(0.03, 0.08))
        url = f"https://api.hh.ru/vacancies/{vacancy_id}"
        
        for attempt in range(3):
            try:
                async with self.session.get(url) as response:
                    if response.status == 200:
                        return await response.json()
                    elif response.status == 429:
                        await asyncio.sleep((attempt + 1) * 1)
                    else:
                        return None
            except:
                if attempt == 2:
                    return None
                await asyncio.sleep(0.3)
        return None
    
    async def process_vacancy(self, vacancy_id):
        if not self.is_running:
            return None
        details = await self.get_vacancy_details(vacancy_id)
        if not details:
            self.processed += 1
            return None
        
        text_parts = []
        if details.get('name'):
            text_parts.append(details['name'])
        if details.get('key_skills'):
            for skill in details['key_skills']:
                if skill.get('name'):
                    text_parts.append(skill['name'])
        if details.get('description'):
            text_parts.append(details['description'])
        snippet = details.get('snippet', {})
        if snippet.get('requirement'):
            text_parts.append(snippet['requirement'])
        if snippet.get('responsibility'):
            text_parts.append(snippet['responsibility'])
        
        full_text = ' '.join(text_parts)
        keywords = extract_keywords(full_text, self.job_titles)
        self.all_keywords.extend(keywords)
        self.processed += 1
        
        if self.processed % 10 == 0:
            progress = (self.processed / self.total) * 100
            self.progress_bar.progress(min(progress / 100, 1.0))
            self.status_text.text(f"⚡ Обработано: {self.processed}/{self.total} ({int(progress)}%)")
            
        return keywords
    
    async def parse_vacancies(self, job_titles, max_vacancies):
        await self.init_session()
        try:
            self.log("🚀 Начинаем сбор ID вакансий...")
            all_ids = []
            vacancies_per_job = max_vacancies // len(job_titles) if job_titles else max_vacancies
            
            for job_title in job_titles:
                if not self.is_running:
                    break
                ids = await self.get_vacancy_ids(job_title, vacancies_per_job)
                all_ids.extend(ids)
                if len(all_ids) >= max_vacancies:
                    all_ids = all_ids[:max_vacancies]
                    break
            
            self.total = len(all_ids)
            self.log(f"📊 Всего ID: {self.total}. Начинаем детальный анализ...")
            
            if self.total == 0:
                return [], Counter(), Counter()
            
            semaphore = asyncio.Semaphore(20)
            async def bounded_process(vid):
                async with semaphore:
                    return await self.process_vacancy(vid)
            
            tasks = [bounded_process(vid) for vid in all_ids]
            await asyncio.gather(*tasks)
            
            keyword_counts = Counter(self.all_keywords)
            sorted_keywords = sorted(keyword_counts.items(), key=lambda x: x[1], reverse=True)
            
            tech_counter = Counter({k: v for k, v in keyword_counts.items() if k in config.all_tech_terms})
            action_counter = Counter({k: v for k, v in keyword_counts.items() if k in config.action_verbs})
            
            self.log(f"📊 Найдено ключевых слов: {len(self.all_keywords)}")
            self.log(f"📊 Уникальных: {len(keyword_counts)}")
            
            return sorted_keywords, tech_counter, action_counter
            
        finally:
            await self.close_session()

# ====================== UI ======================
def main():
    # Инициализация состояния
    if 'results' not in st.session_state:
        st.session_state.results = None
    if 'config_generated' not in st.session_state:
        st.session_state.config_generated = False
    if 'is_running' not in st.session_state:
        st.session_state.is_running = True
    
    # Получаем API ключ из переменных окружения
    api_key = os.getenv('OPENROUTER_API_KEY')
    
    if not api_key:
        st.error("❌ API ключ не найден в переменных окружения")
        st.info("Пожалуйста, обратитесь к администратору для настройки API ключа")
        st.stop()
    
    # Боковая панель
    with st.sidebar:
        st.markdown("---")
        st.markdown(f"""
        <div style="text-align: center; padding: 10px; color: #666;">
            <small>🏠 {os.getenv('APP_URL', 'https://dashkeys.streamlit.app/')}</small>
        </div>
        """, unsafe_allow_html=True)
        
        st.markdown("## ⚙️ Настройки")
        st.info(
            "Как начать:\n"
            "1) Введите одну должность.\n"
            "2) Нажмите 'Сгенерировать конфиг'.\n"
            "3) Нажмите 'Запустить анализ'.\n"
            "4) Дождитесь завершения анализа.\n"
            "5) Изучите вкладки и скачайте Excel."
        )
        
        st.markdown("### 📝 Должность для анализа")
        jobs_input = st.text_area(
            "Введите должность",
            value="Python разработчик",
            height=150,
            help="За один запуск анализируется только одна должность"
        )
        client_ip = get_client_ip()
        
        if jobs_input.strip():
            first_job = jobs_input.strip().split('\n')[0].strip()
            used_today, remaining_today = get_generation_usage_today(client_ip)
            can_generate_now = can_generate_config_today(client_ip)
            if client_ip:
                if can_generate_now:
                    st.caption(
                        f"Генерация конфига для вашего IP: {used_today}/{MAX_GENERATIONS_PER_DAY} за сегодня "
                        f"(осталось {remaining_today})."
                    )
                else:
                    st.warning(
                        f"Для вашего IP дневной лимит исчерпан: {MAX_GENERATIONS_PER_DAY}/{MAX_GENERATIONS_PER_DAY}. "
                        "Попробуйте завтра."
                    )
            else:
                st.caption("IP клиента не определен: ограничение 2 генерации в день не применяется.")
            
            if st.button(f"🎯 Сгенерировать конфиг для '{first_job}'", use_container_width=True):
                if not can_generate_now:
                    st.error("⛔ Лимит исчерпан: генерация конфига доступна 2 раза в день на IP.")
                    st.stop()

                with st.spinner("⏳ анализируем рынок..."):
                    try:
                        job_titles_list = [first_job]
                        new_config = generate_config_from_api(job_titles_list, api_key)
                        config.apply_api_response(new_config, first_job)
                        mark_config_generated_today(client_ip)
                        st.session_state.config_generated = True
                        st.success(f"✅ Конфиг сгенерирован!")
                        st.rerun()
                    except Exception as e:
                        st.error(f"❌ Ошибка: {str(e)}")
                        st.info("Используются дефолтные настройки")
        
        if config.is_generated:
            st.markdown(f"""
            <div class="success-box">
                <b>✓ Активная конфигурация:</b> {config.job_title}<br>
                Категорий: {len(config.tech_categories)} | 
                Технологий: {len(config.all_tech_terms)}
            </div>
            """, unsafe_allow_html=True)
        
        max_vacancies = 120
        min_frequency = 5
        
        st.markdown("---")
        start_btn = st.button("🚀 Запустить анализ", use_container_width=True, type="primary")
        stop_btn = st.button("⏹ Остановить", use_container_width=True)
        
        with st.expander("🔍 Показать текущую конфигурацию"):
            st.json({
                "TECH_CATEGORIES": {k: v[:3] + ["..."] if len(v) > 3 else v 
                                   for k, v in config.tech_categories.items()},
                "ACTION_VERBS (sample)": list(config.action_verbs)[:5]
            })
    
    # Основная область
    st.markdown('<p class="main-header">Аналитический Дашборд</p>', unsafe_allow_html=True)
    
    if not config.is_generated:
        st.info(
            "Быстрый старт:\n"
            "1) Введите одну должность слева.\n"
            "2) Нажмите 'Сгенерировать конфиг'.\n"
            "3) Запустите анализ кнопкой 'Запустить анализ'.\n"
            "4) После расчета откройте вкладки с результатами.\n"
            "5) При необходимости экспортируйте отчет в Excel."
        )
        st.info("💡 Нажмите 'Сгенерировать конфиг' для адаптации под вашу должность или используйте стандартные настройки.")
    
    # Прогресс и статус
    col1, col2, col3 = st.columns([2, 1, 1])
    with col1:
        progress_bar = st.progress(0)
    with col2:
        status_text = st.empty()
    with col3:
        timer_text = st.empty()
    
    with st.expander("📝 Лог выполнения", expanded=False):
        log_container = st.empty()
    
    if stop_btn:
        st.session_state.is_running = False
        st.warning("⏹ Остановка...")
    
    if start_btn:
        job_titles = [j.strip() for j in jobs_input.split('\n') if j.strip()]
        if not job_titles:
            st.error("❌ Введите хотя бы одну должность")
            return
        
        start_time = time.time()
        st.session_state.is_running = True
        
        parser = AsyncHHParser(progress_bar, status_text, log_container, job_titles)
        
        with st.spinner("⏳ Идет сбор данных..."):
            try:
                loop = asyncio.new_event_loop()
                asyncio.set_event_loop(loop)
                results = loop.run_until_complete(
                    parser.parse_vacancies(job_titles, max_vacancies)
                )
                loop.close()
                
                elapsed = time.time() - start_time
                st.session_state.results = results
                st.session_state.elapsed = elapsed
                st.session_state.total_vacancies = parser.processed
                st.session_state.job_titles = job_titles
                
                status_text.text("✅ Готово!")
                progress_bar.progress(1.0)
                logging.info(f"Анализ завершен за {elapsed:.1f}秒. Обработано {parser.processed} вакансий")
                
            except Exception as e:
                st.error(f"❌ Ошибка: {e}")
                logging.error(f"Ошибка при анализе: {e}")
                import traceback
                st.error(traceback.format_exc())
                return
    
    # Отображение результатов
    if st.session_state.results:
        keywords, tech_counter, action_counter = st.session_state.results
        total_vacancies = st.session_state.total_vacancies
        elapsed = st.session_state.elapsed
        job_titles = st.session_state.job_titles
        
        if not keywords:
            st.warning("⚠️ Не найдено ключевых слов. Попробуйте:")
            st.info("1. Увеличить количество вакансий\n2. Изменить должность\n3. Сгенерировать новый конфиг")
            return
        
        max_keyword_freq = keywords[0][1] if keywords else 1
        max_tech_freq = max(tech_counter.values()) if tech_counter else 1
        max_action_freq = max(action_counter.values()) if action_counter else 1
        
        # KPI
        st.markdown("## 📊 Ключевые метрики")
        kpi_col1, kpi_col2, kpi_col3, kpi_col4 = st.columns(4)
        
        with kpi_col1:
            st.metric("⏱ Время", f"{elapsed:.1f}с")
        with kpi_col2:
            st.metric("📄 Вакансий", total_vacancies)
        with kpi_col3:
            st.metric("🔑 Уникальных навыков", len(keywords))
        with kpi_col4:
            top_skill = keywords[0][0] if keywords else "-"
            top_count = keywords[0][1] if keywords else 0
            stars, _, _ = calculate_rating(top_count, max_keyword_freq)
            st.metric("🏆 Топ навык", f"{top_skill} ({stars}★)")
        
        filtered_keywords = [(k, v) for k, v in keywords if v >= min_frequency]
        
        if not filtered_keywords:
            st.warning("⚠️ После фильтрации не осталось данных. Уменьшите минимальную частоту.")
            return
        
        max_filtered_freq = filtered_keywords[0][1]
        
        # ТАБЫ
        tab1, tab2, tab3, tab4, tab5 = st.tabs([
            "🔥 Все ключевые слова", 
            "💻 Технологии", 
            "📂 Категории",
            "⚡ Действия", 
            "📈 Визуализация"
        ])
        
        # === ВКЛАДКА 1 ===
        with tab1:
            col_wordcloud, _ = st.columns([1, 1])
            
            with col_wordcloud:
                st.markdown("### ☁️ Облако слов")
                if len(filtered_keywords) >= 5:
                    word_freq = {k: v for k, v in filtered_keywords[:50]}
                    wc = WordCloud(
                        width=600, height=400, 
                        background_color='#0e1117',
                        colormap='viridis',
                        max_words=50
                    ).generate_from_frequencies(word_freq)
                    
                    fig, ax = plt.subplots(figsize=(8, 5))
                    ax.imshow(wc, interpolation='bilinear')
                    ax.axis('off')
                    st.pyplot(fig)
                else:
                    st.info("Недостаточно данных для облака слов")
            
            st.markdown("### 📋 Топ ключевых слов")
            data_all = []
            for kw, count in filtered_keywords[:100]:
                stars, star_text, pct = calculate_rating(count, max_filtered_freq)
                data_all.append({
                    'Ключевое слово': kw,
                    'Частота': count,
                    'Процент': f"{pct}%",
                    'Рейтинг': star_text
                })
            
            df_all = pd.DataFrame(data_all)
            st.dataframe(df_all, use_container_width=True, hide_index=True)
        
        # === ВКЛАДКА 2 ===
        with tab2:
            if tech_counter:
                st.markdown("### 📊 Топ-15 технологий")
                top_tech = sorted(tech_counter.items(), key=lambda x: x[1], reverse=True)[:15]
                
                if top_tech:
                    fig_tech = px.bar(
                        x=[t[0] for t in top_tech],
                        y=[t[1] for t in top_tech],
                        color=[t[1] for t in top_tech],
                        color_continuous_scale='Viridis',
                        template='plotly_dark',
                        title='Топ-15 технологий по частоте упоминаний'
                    )
                    fig_tech.update_layout(
                        height=400, 
                        title_font_color='white', 
                        font_color='white',
                        xaxis_title="Технологии",
                        yaxis_title="Частота упоминаний"
                    )
                    st.plotly_chart(fig_tech, use_container_width=True)
                
                st.markdown("### 📋 Все технологии")
                tech_data = []
                for tech, count in sorted(tech_counter.items(), key=lambda x: x[1], reverse=True):
                    stars, star_text, pct = calculate_rating(count, max_tech_freq)
                    tech_data.append({
                        'Технология': tech,
                        'Частота': count,
                        'Процент': f"{pct}%",
                        'Рейтинг': star_text
                    })
                
                df_tech = pd.DataFrame(tech_data)
                st.dataframe(df_tech, use_container_width=True, hide_index=True)
            else:
                st.info("Технологии не найдены в вакансиях")
        
        # === ВКЛАДКА 3 ===
        with tab3:
            cat_stats = {}
            for cat, terms in config.tech_categories.items():
                cat_count = sum(tech_counter.get(term, 0) for term in terms)
                if cat_count > 0:
                    cat_stats[cat] = cat_count
            
            if cat_stats:
                col_chart, col_legend = st.columns([3, 2])
                
                with col_chart:
                    st.markdown("### 📊 Распределение по категориям")
                    
                    df_cats = pd.DataFrame({
                        'Категория': list(cat_stats.keys()),
                        'Количество': list(cat_stats.values())
                    })
                    total_cat = df_cats['Количество'].sum()
                    df_cats['Процент'] = (df_cats['Количество'] / total_cat * 100).round(1)
                    
                    colors_for_pie = config.category_colors[:len(df_cats)]
                    
                    fig_pie = go.Figure(data=[go.Pie(
                        labels=df_cats['Категория'],
                        values=df_cats['Количество'],
                        hole=0.4,
                        marker=dict(colors=colors_for_pie, line=dict(color='#0e1117', width=2)),
                        textinfo='percent',
                        textfont=dict(size=14, color='white', family='Arial Black'),
                        textposition='inside'
                    )])
                    
                    fig_pie.update_layout(height=500, showlegend=False, paper_bgcolor='#0e1117', font_color='white')
                    st.plotly_chart(fig_pie, use_container_width=True)
                
                with col_legend:
                    st.markdown("### 🎨 Легенда категорий")
                    
                    for i, (cat, count) in enumerate(sorted(cat_stats.items(), key=lambda x: x[1], reverse=True)):
                        color = config.category_colors[i % len(config.category_colors)]
                        pct = round(count / total_cat * 100, 1)
                        icon = config.category_icons.get(cat, '🔹')
                        
                        cat_terms = {term: tech_counter.get(term, 0) for term in config.tech_categories.get(cat, []) 
                                    if tech_counter.get(term, 0) > 0}
                        sorted_terms = sorted(cat_terms.items(), key=lambda x: x[1], reverse=True)
                        max_cat_freq = max(cat_terms.values()) if cat_terms else 1
                        
                        category_header = f"""
                        <div class="category-card" style="border-left-color: {color}; margin-bottom: 5px;">
                            <div class="category-icon">{icon}</div>
                            <div class="category-color-box" style="background-color: {color};"></div>
                            <div class="category-info">
                                <div class="category-title">{cat}</div>
                                <div class="category-stats">
                                    <span style="color: {color}; font-weight: bold; font-size: 14px;">{count}</span> упоминаний 
                                    <span class="category-percent" style="background: {color}30; color: {color};">
                                        {pct}%
                                    </span>
                                </div>
                            </div>
                        </div>
                        """
                        st.markdown(category_header, unsafe_allow_html=True)
                        
                        if sorted_terms:
                            with st.expander(f"📋 Показать {len(sorted_terms)} технологий"):
                                cols = st.columns(3)
                                for j, (term, count) in enumerate(sorted_terms):
                                    stars, _, pct = calculate_rating(count, max_cat_freq)
                                    with cols[j % 3]:
                                        st.markdown(f"""
                                        <div style="
                                            padding: 8px; 
                                            background-color: #1a1f2e; 
                                            border-radius: 6px; 
                                            margin-bottom: 6px;
                                            border-left: 2px solid {color};
                                        ">
                                            <div style="font-weight: bold; color: #e0e0e0; font-size: 13px;">{term}</div>
                                            <div style="font-size: 11px; color: #a0a0a0;">
                                                {count} раз ({pct}%) {render_stars(stars)}
                                            </div>
                                        </div>
                                        """, unsafe_allow_html=True)
            else:
                st.info("Нет данных по категориям")
        
        # === ВКЛАДКА 4 ===
        with tab4:
            if action_counter:
                col_chart, col_table = st.columns([2, 3])
                
                with col_chart:
                    st.markdown("### 📊 Топ-10 действий")
                    top_actions = sorted(action_counter.items(), key=lambda x: x[1], reverse=True)[:10]
                    
                    if top_actions:
                        fig_actions = px.bar(
                            y=[a[0] for a in top_actions],
                            x=[a[1] for a in top_actions],
                            orientation='h',
                            color=[a[1] for a in top_actions],
                            color_continuous_scale='Blues',
                            template='plotly_dark',
                            title='Топ-10 действий по частоте упоминаний'
                        )
                        fig_actions.update_layout(
                            height=500, 
                            yaxis=dict(autorange="reversed"), 
                            font_color='white',
                            xaxis_title="Частота упоминаний",
                            yaxis_title="Действия"
                        )
                        st.plotly_chart(fig_actions, use_container_width=True)
                
                with col_table:
                    st.markdown("### 📋 Все действия")
                    action_data = []
                    sorted_actions = sorted(action_counter.items(), key=lambda x: x[1], reverse=True)
                    for action, count in sorted_actions:
                        stars, star_text, pct = calculate_rating(count, max_action_freq)
                        action_data.append({
                            'Действие': action,
                            'Частота': count,
                            'Процент': f"{pct}%",
                            'Рейтинг': star_text
                        })
                    
                    df_action = pd.DataFrame(action_data)
                    st.dataframe(df_action, use_container_width=True, hide_index=True)
            else:
                st.info("Действия не найдены")
        
        # === ВКЛАДКА 5 ===
        with tab5:
            st.markdown("## 🎨 Легенда категорий технологий")
            
            cat_stats = {}
            for cat, terms in config.tech_categories.items():
                cat_count = sum(tech_counter.get(term, 0) for term in terms)
                if cat_count > 0:
                    cat_stats[cat] = cat_count
            
            if cat_stats:
                total_cat = sum(cat_stats.values())
                
                cols = st.columns(2)
                for i, (cat, count) in enumerate(sorted(cat_stats.items(), key=lambda x: x[1], reverse=True)):
                    color = config.category_colors[i % len(config.category_colors)]
                    pct = round(count / total_cat * 100, 1)
                    icon = config.category_icons.get(cat, '🔹')
                    
                    cat_terms = {term: tech_counter.get(term, 0) for term in config.tech_categories.get(cat, []) 
                                if tech_counter.get(term, 0) > 0}
                    sorted_terms = sorted(cat_terms.items(), key=lambda x: x[1], reverse=True)[:3]
                    
                    top_terms_parts = []
                    for term, cnt in sorted_terms:
                        top_terms_parts.append(f"{term} ({cnt})")
                    top_terms_str = ", ".join(top_terms_parts)
                    
                    with cols[i % 2]:
                        st.markdown(f"""
                        <div class="category-card" style="border-left-color: {color}; margin-right: 10px;">
                            <div class="category-icon">{icon}</div>
                            <div class="category-color-box" style="background-color: {color};"></div>
                            <div class="category-info">
                                <div class="category-title">{cat}</div>
                                <div class="category-stats">
                                    <span style="color: {color}; font-weight: bold; font-size: 18px;">{count}</span> упоминаний
                                    <div style="margin-top: 5px;">
                                        <div style="background-color: #2a2f3e; height: 8px; border-radius: 4px; overflow: hidden;">
                                            <div style="width: {pct}%; background-color: {color}; height: 100%; border-radius: 4px;"></div>
                                        </div>
                                        <div style="margin-top: 8px; font-size: 12px;">
                                            <span style="color: #a0a0a0;">🏆 Топ-3:</span> 
                                            <span style="color: #e0e0e0;">{top_terms_str if top_terms_str else 'нет данных'}</span>
                                        </div>
                                    </div>
                                </div>
                            </div>
                        </div>
                        """, unsafe_allow_html=True)
                
                st.markdown("---")
                st.markdown("### 📊 Полная детализация по категориям")
                
                for i, (cat, terms) in enumerate(config.tech_categories.items()):
                    cat_terms = {term: tech_counter.get(term, 0) for term in terms if tech_counter.get(term, 0) > 0}
                    if cat_terms:
                        with st.expander(f"{config.category_icons.get(cat, '🔹')} {cat} ({len(cat_terms)} технологий)"):
                            sorted_terms = sorted(cat_terms.items(), key=lambda x: x[1], reverse=True)
                            max_cat_freq = max(cat_terms.values()) if cat_terms else 1
                            
                            cols_detail = st.columns(3)
                            for j, (term, count) in enumerate(sorted_terms):
                                stars, _, pct = calculate_rating(count, max_cat_freq)
                                with cols_detail[j % 3]:
                                    st.markdown(f"""
                                    <div style="
                                        padding: 10px; 
                                        background-color: #1a1f2e; 
                                        border-radius: 8px; 
                                        margin-bottom: 8px; 
                                        border-left: 3px solid {config.category_colors[i % len(config.category_colors)]};
                                    ">
                                        <div style="font-weight: bold; color: #e0e0e0;">{term}</div>
                                        <div style="font-size: 12px; color: #a0a0a0;">
                                            {count} раз ({pct}%) {render_stars(stars)}
                                        </div>
                                    </div>
                                    """, unsafe_allow_html=True)
            else:
                st.info("Нет данных по категориям для отображения")
        
        # === ЭКСПОРТ ===
        st.markdown("---")
        st.markdown("## 💾 Экспорт результатов")
        
        if st.button("📥 Сгенерировать Excel", type="secondary"):
            with st.spinner("Форматирование файла..."):
                data_all = []
                for kw, count in filtered_keywords:
                    stars, star_text, pct = calculate_rating(count, max_filtered_freq)
                    data_all.append([kw, count, star_text])
                df_export_all = pd.DataFrame(data_all, columns=['Ключевое слово', 'Частота', 'Рейтинг'])
                
                tech_export = []
                if tech_counter:
                    for tech, count in sorted(tech_counter.items(), key=lambda x: x[1], reverse=True):
                        stars, star_text, pct = calculate_rating(count, max_tech_freq)
                        tech_export.append([tech, count, star_text])
                df_export_tech = pd.DataFrame(tech_export, columns=['Технология', 'Частота', 'Рейтинг']) if tech_export else pd.DataFrame()
                
                action_export = []
                if action_counter:
                    for action, count in sorted(action_counter.items(), key=lambda x: x[1], reverse=True):
                        stars, star_text, pct = calculate_rating(count, max_action_freq)
                        action_export.append([action, count, star_text])
                df_export_action = pd.DataFrame(action_export, columns=['Действие', 'Частота', 'Рейтинг']) if action_export else pd.DataFrame()
                
                cat_export = []
                for category, terms in config.tech_categories.items():
                    cat_terms = {term: tech_counter.get(term, 0) for term in terms if tech_counter.get(term, 0) > 0}
                    if cat_terms:
                        max_cat = max(cat_terms.values())
                        for term, cnt in sorted(cat_terms.items(), key=lambda x: x[1], reverse=True):
                            stars, star_text, pct = calculate_rating(cnt, max_cat)
                            cat_export.append([category, term, cnt, star_text])
                df_export_cat = pd.DataFrame(cat_export, columns=['Категория', 'Технология', 'Частота', 'Рейтинг']) if cat_export else pd.DataFrame()
                
                dataframes = {
                    'Все ключевые слова': df_export_all,
                    'Технологии': df_export_tech,
                    'Действия': df_export_action,
                    'По категориям': df_export_cat
                }
                
                dataframes = {k: v for k, v in dataframes.items() if not v.empty}
                
                if dataframes:
                    excel_file = create_excel_with_formatting(dataframes, total_vacancies)
                    
                    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                    st.download_button(
                        label="⬇️ Скачать Excel",
                        data=excel_file,
                        file_name=f"info_dashboard_{timestamp}.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                    )
                else:
                    st.warning("Нет данных для экспорта")

if __name__ == "__main__":
    main()